import io
import json
import os
import re

import cv2
import numpy as np
import qrcode
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, Flowable
from reportlab.lib.enums import TA_CENTER

OMR_BUBBLE_RADIUS_MM = 2.0
OMR_BUBBLE_STEP_MM = 14.0
OMR_EDGE_MARGIN_MM = 6.5
OMR_MARKER_SIZE_MM = 2.4
# Vertical offset between guide line and bubble centers in the printed layout.
OMR_GUIDE_TO_BUBBLE_Y_OFFSET_MM = 2.8
# Ignore header/top artifacts and only start OMR reading below this Y ratio.
OMR_ANSWER_START_Y_NORM = 0.22


def _answer_region_min_y(image_height):
    return int(round(image_height * OMR_ANSWER_START_Y_NORM))


class BubbleOptionsRow(Flowable):
    """Draw OMR-friendly bubble options as real circles followed by labels."""

    def __init__(self, options, font_name="Helvetica", font_size=10, left_indent=10):
        super().__init__()
        self.options = options
        self.font_name = font_name
        self.font_size = font_size
        self.left_indent = left_indent
        self.bubble_radius = OMR_BUBBLE_RADIUS_MM * mm
        self.option_step = OMR_BUBBLE_STEP_MM * mm
        self.height = 12 * mm

    def wrap(self, avail_width, avail_height):
        self.width = avail_width
        return avail_width, self.height

    def draw(self):
        c = self.canv
        c.saveState()
        c.setFont(self.font_name, self.font_size - 1)
        c.setLineWidth(1)

        bubble_y = self.height - (4.5 * mm)
        guide_y = bubble_y - (OMR_GUIDE_TO_BUBBLE_Y_OFFSET_MM * mm)
        first_cx = self.left_indent + (4 * mm)
        for i, opt in enumerate(self.options):
            cx = first_cx + (i * self.option_step)
            if cx + self.bubble_radius > self.width:
                break
            c.circle(cx, bubble_y, self.bubble_radius, stroke=1, fill=0)
            c.drawCentredString(cx, bubble_y - (5.5 * mm), str(opt))

        if self.options:
            last_cx = first_cx + ((len(self.options) - 1) * self.option_step)
            line_start = max(self.left_indent, first_cx - (OMR_EDGE_MARGIN_MM * mm))
            line_end = min(self.width - self.left_indent, last_cx + (OMR_EDGE_MARGIN_MM * mm))
            c.setLineWidth(0.8)
            c.line(line_start, guide_y, line_end, guide_y)
            marker_size = OMR_MARKER_SIZE_MM * mm
            c.setFillGray(0)
            c.rect(line_start - (marker_size / 2), guide_y - (marker_size / 2), marker_size, marker_size, stroke=0, fill=1)
            c.rect(line_end - (marker_size / 2), guide_y - (marker_size / 2), marker_size, marker_size, stroke=0, fill=1)

        c.restoreState()


class AnswerRegionStartMarker(Flowable):
    """Visual marker that separates header from answer area in the printed sheet."""

    def __init__(self, label="INICIO DAS RESPOSTAS"):
        super().__init__()
        self.label = label
        self.height = 7 * mm

    def wrap(self, avail_width, avail_height):
        self.width = avail_width
        return avail_width, self.height

    def draw(self):
        c = self.canv
        c.saveState()
        y = self.height - (3 * mm)
        c.setStrokeGray(0.35)
        c.setDash(3, 2)
        c.setLineWidth(0.8)
        c.line(0, y, self.width, y)
        c.setDash()
        c.setFillGray(0.25)
        c.setFont("Helvetica", 7)
        c.drawCentredString(self.width / 2.0, y + (0.8 * mm), self.label)
        c.restoreState()


def generate_test_pdf(test, output_path):
    """Generate a printable PDF for the given test."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=18)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], alignment=TA_CENTER, fontSize=11)
    question_style = ParagraphStyle("Question", parent=styles["Normal"], fontSize=11, spaceAfter=4)

    elements = []

    # Title
    elements.append(Paragraph(test.title, title_style))
    if test.description:
        elements.append(Paragraph(test.description, subtitle_style))
    elements.append(Spacer(1, 6 * mm))

    # Header table: Name + QR code
    template = build_omr_template(test)
    qr_img_buf = _make_qr(f"OMR_TEMPLATE:{template['template_id']}")
    qr_rl = Image(qr_img_buf, width=25 * mm, height=25 * mm)

    header_data = [
        [Paragraph("<b>Nome:</b> ___________________________________", styles["Normal"]),
         Paragraph("<b>Data:</b> _______________", styles["Normal"]),
         qr_rl],
        [Paragraph("<b>Observação:</b> ________________________________________________", styles["Normal"]),
         "", ""],
    ]
    header_table = Table(
        header_data,
        colWidths=[90 * mm, 50 * mm, 30 * mm],
        rowHeights=[12 * mm, 10 * mm],
    )
    header_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
        ("GRID", (0, 0), (-1, 0), 0.5, colors.black),
        ("SPAN", (2, 0), (2, 1)),
        ("SPAN", (0, 1), (1, 1)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(header_table)
    elements.append(AnswerRegionStartMarker())
    elements.append(Spacer(1, 6 * mm))

    # Questions
    for idx, tq in enumerate(test.questions, start=1):
        q = tq.question
        elements.append(Paragraph(f"<b>{idx}.</b> {q.text}", question_style))
        options = q.get_options()
        if options:
            elements.append(BubbleOptionsRow(options))
        elements.append(Spacer(1, 4 * mm))

    doc.build(elements, onFirstPage=lambda canvas, doc_ref: _draw_fiducials(canvas, doc_ref, template))


def _make_qr(data: str) -> io.BytesIO:
    qr = qrcode.QRCode(box_size=4, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


# Minimum fill ratio (0–1) to consider a bubble as marked.
# 40% fill accounts for partial or imprecise shading while avoiding false positives.
OMR_FILL_THRESHOLD = 0.4


def extract_scan_metadata(image_path):
    """
    Try to extract respondent metadata (name/observation) from the scan header.
    OCR is optional; if unavailable, returns empty fields.
    """
    img, source_method = _read_scan_image(image_path)
    if img is None:
        return {"respondent_name": None, "observation": None, "method": source_method}

    try:
        import pytesseract
    except Exception:
        return {"respondent_name": None, "observation": None, "method": "ocr_unavailable"}

    aligned = _align_with_fiducials(img)
    if aligned is not None:
        img = aligned

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    boosted = cv2.convertScaleAbs(gray, alpha=1.5, beta=0)
    cleaned = cv2.GaussianBlur(boosted, (3, 3), 0)
    thresh = cv2.adaptiveThreshold(
        cleaned, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 11
    )

    text = ""
    try:
        text = pytesseract.image_to_string(thresh, lang="por+eng")
    except Exception:
        try:
            text = pytesseract.image_to_string(thresh, lang="eng")
        except Exception:
            return {"respondent_name": None, "observation": None, "method": "ocr_failed"}

    respondent_name = _extract_labeled_text(
        text,
        label_patterns=[r"nome", r"name"],
        stop_patterns=[r"data", r"date", r"observa[çc][aã]o", r"observation"],
    )
    observation = _extract_labeled_text(
        text,
        label_patterns=[r"observa[çc][aã]o", r"observation"],
        stop_patterns=[r"nome", r"name", r"data", r"date"],
    )

    # Fallback OCR on likely header regions when global OCR fails.
    if not respondent_name or not observation:
        region_name, region_obs = _extract_metadata_from_header_regions(thresh)
        respondent_name = respondent_name or region_name
        observation = observation or region_obs

    return {
        "respondent_name": respondent_name,
        "observation": observation,
        "method": f"ocr_tesseract_{source_method}",
    }


def _extract_metadata_from_header_regions(bin_img):
    """OCR header regions with targeted parsing for name/observation."""
    h, w = bin_img.shape[:2]
    if h < 200 or w < 200:
        return None, None

    # Region proportions tuned to generated PDF header area after A4 alignment.
    name_region = bin_img[int(h * 0.08):int(h * 0.18), int(w * 0.06):int(w * 0.72)]
    obs_region = bin_img[int(h * 0.16):int(h * 0.28), int(w * 0.06):int(w * 0.72)]

    try:
        import pytesseract
    except Exception:
        return None, None

    cfg = "--psm 6"
    try:
        name_text = pytesseract.image_to_string(name_region, lang="por+eng", config=cfg)
    except Exception:
        name_text = ""
    try:
        obs_text = pytesseract.image_to_string(obs_region, lang="por+eng", config=cfg)
    except Exception:
        obs_text = ""

    name = _extract_labeled_text(
        name_text,
        label_patterns=[r"nome", r"name"],
        stop_patterns=[r"data", r"date", r"observa[çc][aã]o", r"observation"],
    ) or _sanitize_free_text(name_text)
    obs = _extract_labeled_text(
        obs_text,
        label_patterns=[r"observa[çc][aã]o", r"observation"],
        stop_patterns=[r"nome", r"name", r"data", r"date"],
    ) or _sanitize_free_text(obs_text)
    return name, obs


def _sanitize_free_text(text):
    if not text:
        return None
    cleaned = text.replace("_", " ")
    cleaned = re.sub(r"(?i)\b(nome|name|observa[çc][aã]o|observation|data|date)\b\s*[:\-]?", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .;:-\t\r\n")
    return cleaned or None


def _extract_labeled_text(text, label_patterns, stop_patterns):
    if not text:
        return None

    label_alt = "|".join(label_patterns)
    stop_alt = "|".join(stop_patterns)
    regex = re.compile(
        rf"(?is)\b(?:{label_alt})\b\s*[:\-]?\s*(.+?)(?=(?:\b(?:{stop_alt})\b\s*[:\-]?)|\n|$)"
    )
    match = regex.search(text)
    if not match:
        return None

    value = match.group(1)
    value = value.replace("_", " ").strip(" .;:-\t\r\n")
    value = re.sub(r"\s+", " ", value).strip()
    if not value:
        return None
    return value


def process_scan_image(image_path, test, with_confidence=False):
    """
    Basic OMR: detect filled circles/bubbles in the uploaded scan.
    Returns a dict {question_index (1-based): selected_option_text}.
    Falls back to empty dict if detection is unreliable.
    """
    pages = _read_scan_pages(image_path)
    if not pages:
        return ({}, {}) if with_confidence else {}

    results = {}
    confidences = {}
    remaining_questions = list(test.questions)
    for page_img, _ in pages:
        if not remaining_questions:
            break
        page_results, page_confidences, consumed = _process_scan_page(page_img, remaining_questions)
        results.update(page_results)
        confidences.update(page_confidences)
        if consumed > 0:
            remaining_questions = remaining_questions[consumed:]

    return (results, confidences) if with_confidence else results


def _process_scan_page(img, questions):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    thresh = cv2.adaptiveThreshold(
        blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 31, 8
    )
    aligned = _align_with_fiducials(img)
    if aligned is not None:
        gray = cv2.cvtColor(aligned, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        thresh = cv2.adaptiveThreshold(
            blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 31, 8
        )

    guide_results = _process_by_guide_lines(gray, thresh, questions)
    if guide_results is not None:
        return guide_results

    circles = cv2.HoughCircles(
        blurred,
        cv2.HOUGH_GRADIENT,
        dp=1.2,
        minDist=15,
        param1=50,
        param2=30,
        minRadius=8,
        maxRadius=20,
    )
    if circles is None:
        return {}, {}, 0

    circles = np.round(circles[0, :]).astype("int")
    filtered = []
    h, w = gray.shape[:2]
    min_answer_y = _answer_region_min_y(h)
    for c in circles:
        cx, cy, r = c
        if r < 7 or r > 18:
            continue
        if cx < int(w * 0.05) or cx > int(w * 0.65):
            continue
        if cy < min_answer_y or cy > int(h * 0.96):
            continue
        filtered.append(c)
    circles = filtered
    if not circles:
        return {}, {}, 0

    circles = sorted(circles, key=lambda c: c[1])
    question_count = len(questions)
    rows = _cluster_circles_into_rows(circles, question_count)
    if not rows:
        rows = _group_circles_by_row(circles)
        rows = [sorted(row, key=lambda c: c[0]) for row in rows if len(row) >= 2]
        rows = sorted(rows, key=lambda row: np.mean([c[1] for c in row]))

    results = {}
    confidences = {}
    row_idx = 0
    consumed = 0
    for tq in questions:
        q = tq.question
        options = q.get_options()
        if not options:
            consumed += 1
            continue
        expected_count = len(options)
        q_row = None
        while row_idx < len(rows):
            candidate = rows[row_idx]
            row_idx += 1
            if len(candidate) >= expected_count:
                q_row = candidate
                consumed += 1
                break
        if q_row is None:
            break

        expected_step = float((w / 210.0) * 14.0)
        q_circles = _pick_row_circles(q_row, expected_count, expected_step=expected_step)
        if len(q_circles) < expected_count:
            continue

        best = None
        best_score = -1
        best_fill_inner = -1
        second_best_score = -1
        option_scores = []
        option_inner = []
        for i, (cx, cy, r) in enumerate(q_circles):
            inner_r = max(2, int(r * 0.55))
            inner_mask = np.zeros(gray.shape, dtype="uint8")
            cv2.circle(inner_mask, (cx, cy), inner_r, 255, -1)
            inner_filled = cv2.countNonZero(cv2.bitwise_and(thresh, thresh, mask=inner_mask))
            inner_total = cv2.countNonZero(inner_mask)
            inner_ratio = inner_filled / inner_total if inner_total > 0 else 0

            ring_outer = max(inner_r + 1, int(r * 0.95))
            ring_inner = max(1, int(r * 0.72))
            ring_mask = np.zeros(gray.shape, dtype="uint8")
            cv2.circle(ring_mask, (cx, cy), ring_outer, 255, -1)
            cv2.circle(ring_mask, (cx, cy), ring_inner, 0, -1)
            ring_filled = cv2.countNonZero(cv2.bitwise_and(thresh, thresh, mask=ring_mask))
            ring_total = cv2.countNonZero(ring_mask)
            ring_ratio = ring_filled / ring_total if ring_total > 0 else 0

            score = inner_ratio - (0.30 * ring_ratio)
            option_scores.append(score)
            option_inner.append(inner_ratio)
            if score > best_score:
                second_best_score = best_score
                best_score = score
                best_fill_inner = inner_ratio
                best = options[i]

        selected = _select_marked_option(options, option_scores, option_inner)
        if selected is not None:
            sel_idx, sel_score = selected
            results[tq.question_id] = options[sel_idx]
            confidences[tq.question_id] = round(float(max(0.0, sel_score)), 4)

    return results, confidences, consumed


def _process_by_guide_lines(gray, thresh, test_questions):
    """
    Deterministic OMR using printed horizontal guide lines.
    This is preferred over Hough circles when guide lines are detected.
    """
    h, w = gray.shape[:2]
    marker_rows = _detect_answer_marker_rows(thresh, len(test_questions))
    guide_rows = _detect_answer_guide_rows(thresh, len(test_questions))

    # Prefer the method that detects more rows; this avoids fixed false-positives
    # from marker-like artifacts and prevents repeating the same 1-2 answers.
    if len(guide_rows) > len(marker_rows):
        rows = guide_rows
    else:
        rows = marker_rows if marker_rows else guide_rows

    if len(rows) < 1:
        return None
    if len(rows) < min(3, len(test_questions)):
        # Too few rows is usually noise; let circle-based fallback try this page.
        return None

    px_per_mm = w / 210.0
    bubble_step = int(round(OMR_BUBBLE_STEP_MM * px_per_mm))
    edge_margin = int(round(OMR_EDGE_MARGIN_MM * px_per_mm))
    bubble_radius = max(7, int(round(OMR_BUBBLE_RADIUS_MM * px_per_mm)))
    bubble_y_offset = int(round(OMR_GUIDE_TO_BUBBLE_Y_OFFSET_MM * px_per_mm))

    results = {}
    confidences = {}
    consumed = min(len(rows), len(test_questions))
    for idx, tq in enumerate(test_questions[:consumed]):
        q = tq.question
        options = q.get_options()
        if not options:
            continue
        row = rows[idx]
        line_y = int(row["y"])
        line_start = int(row["x1"])
        line_end = int(row["x2"])
        first_cx = line_start + edge_margin
        last_cx = line_end - edge_margin
        x_positions = []
        if len(options) > 1:
            step = max(1.0, (last_cx - first_cx) / float(len(options) - 1))
            for i in range(len(options)):
                x_positions.append(int(round(first_cx + (i * step))))
        else:
            x_positions.append(int(round((first_cx + last_cx) / 2.0)))
        y = _estimate_bubble_center_y(
            thresh=thresh,
            line_y=line_y,
            x_positions=x_positions,
            bubble_radius=bubble_radius,
            fallback_y=max(0, line_y - bubble_y_offset),
            px_per_mm=px_per_mm,
        )
        x_shift = _estimate_bubble_x_shift(
            thresh=thresh,
            y=y,
            x_positions=x_positions,
            bubble_radius=bubble_radius,
            px_per_mm=px_per_mm,
        )
        if x_shift != 0:
            x_positions = [x + x_shift for x in x_positions]
        best = None
        best_score = -1
        second_best_score = -1
        best_fill_inner = -1
        option_scores = []
        option_inner = []

        for i, opt in enumerate(options):
            cx = x_positions[i] if i < len(x_positions) else None
            if cx is None:
                continue
            if cx <= 0 or cx >= (w - 1):
                continue
            inner_r = max(2, int(bubble_radius * 0.55))
            inner_mask = np.zeros(gray.shape, dtype="uint8")
            cv2.circle(inner_mask, (cx, y), inner_r, 255, -1)
            inner_filled = cv2.countNonZero(cv2.bitwise_and(thresh, thresh, mask=inner_mask))
            inner_total = cv2.countNonZero(inner_mask)
            inner_ratio = inner_filled / inner_total if inner_total > 0 else 0

            ring_outer = max(inner_r + 1, int(bubble_radius * 0.95))
            ring_inner = max(1, int(bubble_radius * 0.72))
            ring_mask = np.zeros(gray.shape, dtype="uint8")
            cv2.circle(ring_mask, (cx, y), ring_outer, 255, -1)
            cv2.circle(ring_mask, (cx, y), ring_inner, 0, -1)
            ring_filled = cv2.countNonZero(cv2.bitwise_and(thresh, thresh, mask=ring_mask))
            ring_total = cv2.countNonZero(ring_mask)
            ring_ratio = ring_filled / ring_total if ring_total > 0 else 0

            score = inner_ratio - (0.30 * ring_ratio)
            option_scores.append(score)
            option_inner.append(inner_ratio)
            if score > best_score:
                second_best_score = best_score
                best_score = score
                best_fill_inner = inner_ratio
                best = opt

        selected = _select_marked_option(options, option_scores, option_inner)
        if selected is not None:
            sel_idx, sel_score = selected
            results[tq.question_id] = options[sel_idx]
            confidences[tq.question_id] = round(float(max(0.0, sel_score)), 4)

    return results, confidences, consumed


def _estimate_bubble_center_y(thresh, line_y, x_positions, bubble_radius, fallback_y, px_per_mm):
    """
    Estimate bubble center Y by scanning near the guide line and maximizing
    average ring response (circle border energy) across expected bubble Xs.
    """
    h, w = thresh.shape[:2]
    if not x_positions:
        return max(0, min(h - 1, fallback_y))

    # Search around the expected center (from template geometry), not around
    # the guide line itself. This avoids locking onto the printed guide line
    # and creating deterministic false positives on blank sheets.
    search_radius = int(round(2.5 * px_per_mm))
    y_min = max(0, fallback_y - search_radius)
    y_max = min(h - 1, fallback_y + search_radius)
    if y_min >= y_max:
        return max(0, min(h - 1, fallback_y))

    probe_positions = [x for x in x_positions if 0 < x < (w - 1)]
    if len(probe_positions) > 5:
        probe_positions = probe_positions[:5]
    if not probe_positions:
        return max(0, min(h - 1, fallback_y))

    ring_outer = max(2, int(bubble_radius * 0.95))
    ring_inner = max(1, int(bubble_radius * 0.72))
    best_y = fallback_y
    best_signal = -1.0
    guide_guard = int(round(1.2 * px_per_mm))
    for y in range(y_min, y_max + 1):
        if abs(y - line_y) <= guide_guard:
            continue
        signal_sum = 0.0
        for cx in probe_positions:
            ring_mask = np.zeros(thresh.shape, dtype="uint8")
            cv2.circle(ring_mask, (cx, y), ring_outer, 255, -1)
            cv2.circle(ring_mask, (cx, y), ring_inner, 0, -1)
            filled = cv2.countNonZero(cv2.bitwise_and(thresh, thresh, mask=ring_mask))
            total = cv2.countNonZero(ring_mask)
            signal_sum += (filled / total) if total > 0 else 0.0
        avg_signal = signal_sum / float(len(probe_positions))
        if avg_signal > best_signal:
            best_signal = avg_signal
            best_y = y

    if best_signal < 0:
        return max(0, min(h - 1, fallback_y))
    return max(0, min(h - 1, best_y))


def _estimate_bubble_x_shift(thresh, y, x_positions, bubble_radius, px_per_mm):
    """
    Estimate horizontal offset for bubble centers in a row.
    Compensates printer/scan stretch and marker pairing inaccuracies.
    """
    h, w = thresh.shape[:2]
    if not x_positions or y <= 0 or y >= (h - 1):
        return 0

    probe_positions = [x for x in x_positions if 0 < x < (w - 1)]
    if len(probe_positions) > 5:
        probe_positions = probe_positions[:5]
    if not probe_positions:
        return 0

    ring_outer = max(2, int(bubble_radius * 0.95))
    ring_inner = max(1, int(bubble_radius * 0.72))
    max_shift = int(round(2.5 * px_per_mm))
    best_shift = 0
    best_signal = -1.0
    signal_by_shift = {}

    for dx in range(-max_shift, max_shift + 1):
        signal_sum = 0.0
        used = 0
        for cx in probe_positions:
            xx = cx + dx
            if xx <= 0 or xx >= (w - 1):
                continue
            ring_mask = np.zeros(thresh.shape, dtype="uint8")
            cv2.circle(ring_mask, (xx, y), ring_outer, 255, -1)
            cv2.circle(ring_mask, (xx, y), ring_inner, 0, -1)
            filled = cv2.countNonZero(cv2.bitwise_and(thresh, thresh, mask=ring_mask))
            total = cv2.countNonZero(ring_mask)
            signal_sum += (filled / total) if total > 0 else 0.0
            used += 1
        if used == 0:
            continue
        avg_signal = signal_sum / float(used)
        signal_by_shift[dx] = avg_signal
        if avg_signal > best_signal:
            best_signal = avg_signal
            best_shift = dx

    base_signal = signal_by_shift.get(0, -1.0)
    if best_shift != 0 and base_signal >= 0 and best_signal < (base_signal + 0.015):
        return 0
    return best_shift


def _select_marked_option(options, scores, inner_ratios):
    """
    Return (index, score) only when evidence is strong enough.
    Prevents deterministic false-positives across different scans.
    """
    if not options or not scores or len(scores) != len(options):
        return None

    ranked = sorted(enumerate(scores), key=lambda it: it[1], reverse=True)
    best_idx, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else -1.0
    best_inner = inner_ratios[best_idx] if best_idx < len(inner_ratios) else 0.0
    second_idx = ranked[1][0] if len(ranked) > 1 else None
    second_inner = inner_ratios[second_idx] if second_idx is not None and second_idx < len(inner_ratios) else 0.0

    mean_score = float(np.mean(scores))
    std_score = float(np.std(scores))
    mean_inner = float(np.mean(inner_ratios)) if inner_ratios else 0.0
    std_inner = float(np.std(inner_ratios)) if inner_ratios else 0.0
    margin_score = best_score - second_score
    margin_inner = best_inner - second_inner

    score_is_outlier = best_score >= (mean_score + max(0.06, 1.25 * std_score))
    inner_is_outlier = best_inner >= (mean_inner + max(0.10, 1.35 * std_inner))
    separation_ok = margin_score >= max(0.08, 1.10 * std_score) and margin_inner >= max(0.07, 1.00 * std_inner)

    # If two options look similarly filled, don't guess.
    ambiguous_multi_mark = second_inner >= 0.22 and margin_inner < 0.06
    if ambiguous_multi_mark:
        return None

    if best_inner >= OMR_FILL_THRESHOLD:
        return best_idx, best_score
    if best_inner >= 0.24 and score_is_outlier and inner_is_outlier and separation_ok:
        return best_idx, best_score
    return None


def _detect_answer_marker_rows(thresh, expected_rows):
    """Detect row markers (small black squares) printed at both ends of each answer line."""
    h, w = thresh.shape[:2]
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    markers = []
    min_answer_y = _answer_region_min_y(h)
    for c in contours:
        x, y, bw, bh = cv2.boundingRect(c)
        rect_area = bw * bh
        if rect_area < 70 or rect_area > 600:
            continue
        ratio = bw / float(bh) if bh else 0
        if ratio < 0.6 or ratio > 1.6:
            continue
        area = cv2.contourArea(c)
        fill_ratio = area / float(rect_area) if rect_area else 0
        if fill_ratio < 0.55:
            continue
        cx = int(x + (bw / 2))
        cy = int(y + (bh / 2))
        if cx < int(w * 0.05) or cx > int(w * 0.60):
            continue
        if cy < min_answer_y or cy > int(h * 0.96):
            continue
        markers.append({"x": cx, "y": cy, "size": rect_area})

    if len(markers) < 2:
        return []

    markers.sort(key=lambda m: m["y"])
    groups = []
    for m in markers:
        target = None
        for g in groups:
            if abs(m["y"] - g["y"]) <= 12:
                target = g
                break
        if target is None:
            groups.append({"y": m["y"], "items": [m]})
        else:
            target["items"].append(m)
            target["y"] = int(round(np.mean([it["y"] for it in target["items"]])))

    rows = []
    for g in groups:
        items = sorted(g["items"], key=lambda it: it["x"])
        if len(items) < 2:
            continue
        left = items[0]
        right = items[-1]
        length = right["x"] - left["x"]
        if length < int(w * 0.20) or length > int(w * 0.48):
            continue
        rows.append({"x1": left["x"], "x2": right["x"], "y": g["y"], "length": length})

    if not rows:
        return []
    rows.sort(key=lambda r: r["y"])
    if len(rows) <= expected_rows:
        return rows
    return rows[:expected_rows]


def _detect_answer_guide_rows(thresh, expected_rows):
    """Detect the horizontal guide rows printed with the bubbles."""
    h, w = thresh.shape[:2]
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(20, w // 30), 1))
    horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel)
    lines = cv2.HoughLinesP(
        horizontal,
        rho=1,
        theta=np.pi / 180,
        threshold=50,
        minLineLength=max(140, int(w * 0.20)),
        maxLineGap=18,
    )
    if lines is None:
        return []

    candidates = []
    min_answer_y = _answer_region_min_y(h)
    for line in lines[:, 0, :]:
        x1, y1, x2, y2 = map(int, line)
        if abs(y1 - y2) > 3:
            continue
        lx1, lx2 = sorted((x1, x2))
        length = lx2 - lx1
        if length < int(w * 0.20) or length > int(w * 0.45):
            continue
        if lx1 < int(w * 0.06) or lx1 > int(w * 0.52):
            continue
        if lx2 > int(w * 0.70):
            continue
        y = int((y1 + y2) / 2)
        if y < min_answer_y or y > int(h * 0.96):
            continue
        candidates.append({"x1": lx1, "x2": lx2, "y": y, "length": length})

    if not candidates:
        return []

    candidates.sort(key=lambda r: r["y"])
    grouped = []
    for row in candidates:
        target = None
        for g in grouped:
            if abs(row["y"] - g["y"]) <= 10:
                target = g
                break
        if target is None:
            grouped.append(dict(row))
        else:
            # Keep the strongest (longest) segment for each row cluster.
            if row["length"] > target["length"]:
                target.update(row)

    grouped.sort(key=lambda r: r["y"])
    if len(grouped) <= expected_rows:
        return grouped

    # Prefer the longest rows and then restore top-to-bottom order.
    top = sorted(grouped, key=lambda r: r["length"], reverse=True)[:expected_rows]
    return sorted(top, key=lambda r: r["y"])


def _cluster_circles_into_rows(circles, question_count):
    """
    Cluster circles into exactly N question rows (deterministic assignment).
    Falls back to heuristic grouping when clustering is not reliable.
    """
    if not circles or question_count <= 0:
        return []
    if len(circles) < question_count * 2:
        return []

    ys = np.array([[float(c[1])] for c in circles], dtype=np.float32)
    k = min(question_count, len(circles))
    if k < 1:
        return []

    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 50, 0.2)
    flags = cv2.KMEANS_PP_CENTERS
    compactness, labels, centers = cv2.kmeans(ys, k, None, criteria, 8, flags)
    _ = compactness
    labels = labels.reshape(-1)
    centers = centers.reshape(-1)

    grouped = {}
    for idx, label in enumerate(labels):
        grouped.setdefault(int(label), []).append(circles[idx])

    rows = []
    for label, row_circles in grouped.items():
        if len(row_circles) < 2:
            continue
        row_sorted = sorted(row_circles, key=lambda c: c[0])
        rows.append((float(centers[label]), row_sorted))

    if len(rows) < question_count:
        return []

    rows = sorted(rows, key=lambda item: item[0])[:question_count]
    return [row for _, row in rows]


def _group_circles_by_row(circles, row_tolerance=24):
    """Group circles by similar Y position to preserve question rows."""
    rows = []
    for circle in circles:
        _, cy, _ = circle
        target_row = None
        best_delta = None
        for row in rows:
            mean_y = np.mean([c[1] for c in row])
            delta = abs(cy - mean_y)
            if delta <= row_tolerance and (best_delta is None or delta < best_delta):
                target_row = row
                best_delta = delta
        if target_row is None:
            rows.append([circle])
        else:
            target_row.append(circle)
    return rows


def _pick_row_circles(row_circles, expected_count, expected_step=None):
    """
    Choose the best contiguous subset of circles within a row.
    Prefers more regular horizontal spacing and larger average radius.
    """
    if len(row_circles) <= expected_count:
        return row_circles

    best_window = row_circles[:expected_count]
    best_score = None
    for i in range(0, len(row_circles) - expected_count + 1):
        window = row_circles[i:i + expected_count]
        xs = [c[0] for c in window]
        rs = [c[2] for c in window]
        if len(xs) > 1:
            spacings = np.diff(xs)
            spacing_std = float(np.std(spacings))
        else:
            spacing_std = 0.0
        radius_mean = float(np.mean(rs))
        mean_spacing = float(np.mean(np.diff(xs))) if len(xs) > 1 else 0.0
        step_penalty = 0.0
        if expected_step and len(xs) > 1:
            step_penalty = abs(mean_spacing - expected_step)
        # Lower score is better: regular spacing near template step and larger circles.
        score = (1.4 * spacing_std) + step_penalty - (0.25 * radius_mean)
        if best_score is None or score < best_score:
            best_score = score
            best_window = window
    return best_window


def _read_scan_image(image_path):
    """Read scan input (image or PDF first page) and return (bgr_image_or_none, method)."""
    ext = os.path.splitext(image_path)[1].lower()
    if ext == ".pdf":
        try:
            import pypdfium2 as pdfium
        except Exception:
            return None, "pdf_renderer_unavailable"
        try:
            pdf = pdfium.PdfDocument(image_path)
            if len(pdf) < 1:
                return None, "pdf_empty"
            page = pdf[0]
            pil_img = page.render(scale=2.0).to_pil()
            arr = np.array(pil_img)
            if arr.ndim == 2:
                bgr = cv2.cvtColor(arr, cv2.COLOR_GRAY2BGR)
            elif arr.shape[2] == 4:
                bgr = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)
            else:
                bgr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
            return bgr, "pdf_page1_rendered"
        except Exception:
            return None, "pdf_render_failed"

    img = cv2.imread(image_path)
    if img is None:
        return None, "image_unreadable"
    return img, "image_loaded"


def _read_scan_pages(image_path):
    """Read scan input and return list of (bgr_image, method) pages."""
    ext = os.path.splitext(image_path)[1].lower()
    if ext == ".pdf":
        try:
            import pypdfium2 as pdfium
        except Exception:
            return []
        pages = []
        try:
            pdf = pdfium.PdfDocument(image_path)
            for idx in range(len(pdf)):
                page = pdf[idx]
                pil_img = page.render(scale=2.0).to_pil()
                arr = np.array(pil_img)
                if arr.ndim == 2:
                    bgr = cv2.cvtColor(arr, cv2.COLOR_GRAY2BGR)
                elif arr.shape[2] == 4:
                    bgr = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)
                else:
                    bgr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
                pages.append((bgr, f"pdf_page{idx + 1}_rendered"))
        except Exception:
            return []
        return pages

    img = cv2.imread(image_path)
    if img is None:
        return []
    return [(img, "image_loaded")]


def import_questions_from_xlsx(file_path):
    """
    Import questions from an Excel file.
    Expected columns: text, answer_type, correct_answer, custom_options
    Returns list of dicts.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data = dict(zip(headers, row))
        q = {
            "text": str(data.get("text") or data.get("pergunta") or "").strip(),
            "answer_type": str(data.get("answer_type") or data.get("tipo") or "yes_no").strip().lower(),
            "correct_answer": str(data.get("correct_answer") or data.get("resposta_correta") or "").strip(),
            "custom_options": str(data.get("custom_options") or data.get("opcoes") or "").strip(),
        }
        if q["text"]:
            if q["answer_type"] not in ("yes_no", "likert", "custom"):
                import warnings
                warnings.warn(
                    f"Tipo de resposta desconhecido '{q['answer_type']}'; usando 'yes_no'.",
                    UserWarning,
                    stacklevel=2,
                )
                q["answer_type"] = "yes_no"
            questions.append(q)
    return questions


def import_questions_from_json(file_obj_or_path):
    """
    Import questions from JSON payload/file.
    Accepted payload:
    {
      "questions": [
        {
          "text": "...",
          "answer_type": "yes_no|likert|custom",
          "correct_answer": "...",
          "custom_options": "A, B, C"
        }
      ]
    }
    """
    if isinstance(file_obj_or_path, str):
        with open(file_obj_or_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
    else:
        raw = file_obj_or_path.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
        payload = json.loads(raw)

    rows = payload.get("questions", [])
    questions = []
    for row in rows:
        q = {
            "text": str(row.get("text") or row.get("pergunta") or "").strip(),
            "answer_type": str(row.get("answer_type") or row.get("tipo") or "yes_no").strip().lower(),
            "correct_answer": str(row.get("correct_answer") or row.get("resposta_correta") or "").strip(),
            "custom_options": str(row.get("custom_options") or row.get("opcoes") or "").strip(),
        }
        if q["text"]:
            questions.append(q)
    return questions


def export_questions_to_xlsx(questions):
    """Return BytesIO with questions exported to xlsx."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perguntas"
    headers = ["text", "answer_type", "correct_answer", "custom_options"]
    ws.append(headers)
    for q in questions:
        ws.append([q.text, q.answer_type, q.correct_answer or "", q.custom_options or ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_omr_template(test):
    """Build a versioned OMR template metadata descriptor for the test."""
    questions = []
    for order, tq in enumerate(test.questions, start=1):
        q = tq.question
        questions.append(
            {
                "order": order,
                "question_id": q.id,
                "answer_type": q.answer_type,
                "options": q.get_options(),
                "option_count": len(q.get_options()),
            }
        )
    return {
        "template_version": "1.1",
        "template_id": f"test-{test.id}-v1_1",
        "fiducials": [
            {"name": "tl", "x_norm": 0.03, "y_norm": 0.03},
            {"name": "tr", "x_norm": 0.97, "y_norm": 0.03},
            {"name": "bl", "x_norm": 0.03, "y_norm": 0.97},
            {"name": "br", "x_norm": 0.97, "y_norm": 0.97},
        ],
        "questions": questions,
    }


def _draw_fiducials(canvas, doc_ref, template):
    width, height = A4
    size = 5 * mm
    for fid in template["fiducials"]:
        x = fid["x_norm"] * width
        y = fid["y_norm"] * height
        canvas.setFillColorRGB(0, 0, 0)
        canvas.rect(x - (size / 2), y - (size / 2), size, size, stroke=0, fill=1)


def _align_with_fiducials(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape[:2]
    thresh = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 31, 8
    )
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    squares = []
    for c in contours:
        area = cv2.contourArea(c)
        if area < 80:
            continue
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.08 * peri, True)
        if len(approx) == 4:
            x, y, w, h = cv2.boundingRect(approx)
            ratio = w / float(h) if h else 0
            if 0.6 <= ratio <= 1.4:
                squares.append((area, x, y, w, h))

    if len(squares) < 4:
        return None
    squares = sorted(squares, key=lambda s: s[0], reverse=True)[:8]
    centers = np.array([[x + w / 2.0, y + h / 2.0] for _, x, y, w, h in squares], dtype=np.float32)
    if centers.shape[0] < 4:
        return None

    s = centers.sum(axis=1)
    diff = np.diff(centers, axis=1).reshape(-1)
    tl = centers[np.argmin(s)]
    br = centers[np.argmax(s)]
    tr = centers[np.argmin(diff)]
    bl = centers[np.argmax(diff)]

    corner_margin_x = w * 0.20
    corner_margin_y = h * 0.20
    if not (tl[0] <= corner_margin_x and tl[1] <= corner_margin_y):
        return None
    if not (tr[0] >= (w - corner_margin_x) and tr[1] <= corner_margin_y):
        return None
    if not (bl[0] <= corner_margin_x and bl[1] >= (h - corner_margin_y)):
        return None
    if not (br[0] >= (w - corner_margin_x) and br[1] >= (h - corner_margin_y)):
        return None

    if (tr[0] - tl[0]) < (w * 0.55) or (bl[1] - tl[1]) < (h * 0.55):
        return None

    src = np.array([tl, tr, br, bl], dtype=np.float32)
    out_w, out_h = 1654, 2339  # ~A4 @ 150 DPI
    dst = np.array([[0, 0], [out_w - 1, 0], [out_w - 1, out_h - 1], [0, out_h - 1]], dtype=np.float32)
    m = cv2.getPerspectiveTransform(src, dst)
    return cv2.warpPerspective(img, m, (out_w, out_h))
